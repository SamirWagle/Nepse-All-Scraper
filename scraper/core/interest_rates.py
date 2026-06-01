"""
Bank interest-rate scraper (Nepal) — NRB monthly statistics, authoritative.

Source: NRB "Banking & Financial Statistics (Monthly)" Excel workbooks, listed at
    https://www.nrb.org.np/category/monthly-statistics/
Each workbook's sheet "C15" ("Overall Summary of Interest Rate on LCY Deposit")
holds the FULL monthly time series across columns — one file backfills the whole
history (Asar 2070 / Mid-Jul 2013 → latest) AND carries the newest month. So the
same parser serves both:
  - backfill(): ingest every monthly column from the latest workbook
  - update():   append only the newest month (idempotent per calendar month)

Series stored: weighted-average FIXED-deposit rate (the "FD rate" the prediction
references), from the workbook's "Fixed:" row.

Data file: data/interest_rates/fd_rates.csv  (columns: date,rate,metric,source)

Fails graceful: any fetch/parse error logs and leaves the CSV untouched so the
daily run is never blocked.
"""
import csv
import io
import logging
import re
from datetime import date
from pathlib import Path

import openpyxl
import requests

logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "interest_rates"
CSV_PATH = DATA_DIR / "fd_rates.csv"
CSV_FIELDS = ["date", "rate", "metric", "source"]

LISTING_URL = "https://www.nrb.org.np/category/monthly-statistics/"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 60

DEPOSIT_SHEET = "C15"
FIXED_ROW_LABEL = "fixed"          # row 2 "Fixed:"
PERIOD_HEADER_ROW = 4              # row holding "Asar, 2070 (Mid Jul, 2013)" labels
SOURCE_TAG = "NRB monthly statistics (xlsx, C15 Fixed deposit Wt.avg)"

MIN_RATE, MAX_RATE = 0.5, 25.0     # plausibility guard

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "jly": 7, "july": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10,
    "nov": 11, "dec": 12,
}
# "... (Mid July, 2025)" / "(Mid Jly, 2018)" → month, year
_PERIOD_RE = re.compile(r"Mid\s+([A-Za-z]+)\.?,?\s*(\d{4})", re.IGNORECASE)
_XLSX_RE = re.compile(r'https://www\.nrb\.org\.np/contents/uploads/[^"\']+?\.xlsx')


def _period_to_date(label):
    """'Ashar, 2082 (Mid July, 2025)' → '2025-07-15'. None if unparseable."""
    if not label:
        return None
    m = _PERIOD_RE.search(str(label))
    if not m:
        return None
    mon = _MONTHS.get(m.group(1).strip().lower()[:4]) or _MONTHS.get(m.group(1).strip().lower()[:3])
    if not mon:
        return None
    return f"{int(m.group(2)):04d}-{mon:02d}-15"


def _fetch_latest_xlsx_url():
    """Scrape the listing page → newest monthly-statistics .xlsx URL, or None."""
    headers = {"User-Agent": USER_AGENT}
    resp = requests.get(LISTING_URL, headers=headers, timeout=REQUEST_TIMEOUT, verify=False)
    resp.raise_for_status()
    urls = _XLSX_RE.findall(resp.text)
    # Skip asset images (logo etc.) — already excluded by .xlsx filter.
    return urls[0] if urls else None


def _download(url):
    headers = {"User-Agent": USER_AGENT}
    resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT, verify=False)
    resp.raise_for_status()
    return resp.content


def parse_deposit_series(xlsx_source):
    """Parse the C15 sheet → [{date, rate}] sorted by date (fixed-deposit Wt.avg).

    xlsx_source: path, bytes, or file-like.
    """
    if isinstance(xlsx_source, (bytes, bytearray)):
        xlsx_source = io.BytesIO(xlsx_source)
    wb = openpyxl.load_workbook(xlsx_source, read_only=True, data_only=True)
    if DEPOSIT_SHEET not in wb.sheetnames:
        logger.warning("Sheet %s not found in workbook.", DEPOSIT_SHEET)
        return []
    ws = wb[DEPOSIT_SHEET]
    rows = {rn: row for rn, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1)}

    header = rows.get(PERIOD_HEADER_ROW, ())
    # Find the "Fixed:" row by its label in column index 2 (S.N.=1, Description=2).
    fixed_row = None
    for row in rows.values():
        label = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
        if label.startswith(FIXED_ROW_LABEL):
            fixed_row = row
            break
    if not header or fixed_row is None:
        logger.warning("Could not locate period header or Fixed row in %s.", DEPOSIT_SHEET)
        return []

    series = {}
    for col, cell in enumerate(header):
        if not cell or "Mid" not in str(cell):
            continue
        d = _period_to_date(cell)
        if not d or col >= len(fixed_row):
            continue
        val = fixed_row[col]
        try:
            rate = round(float(val), 2)
        except (TypeError, ValueError):
            continue
        if not (MIN_RATE <= rate <= MAX_RATE):
            continue
        series[d] = rate  # later periods overwrite earlier dup dates (none expected)

    return [{"date": d, "rate": series[d]} for d in sorted(series)]


# ----------------------------------------------------------------------
# CSV helpers (immutable read; full rewrite on change)
# ----------------------------------------------------------------------
def _read_rows():
    if not CSV_PATH.exists():
        return []
    with open(CSV_PATH, newline="") as f:
        return list(csv.DictReader(f))


def _write_rows(rows):
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(rows, key=lambda r: r.get("date", ""))
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in ordered:
            writer.writerow({k: r.get(k, "") for k in CSV_FIELDS})


def _merge(existing, parsed, source_tag):
    """Merge parsed [{date,rate}] into existing rows. NRB data wins on date clash."""
    by_date = {r["date"]: dict(r) for r in existing}
    for p in parsed:
        by_date[p["date"]] = {
            "date": p["date"],
            "rate": f"{p['rate']:.2f}",
            "metric": "fixed_deposit",
            "source": source_tag,
        }
    return list(by_date.values())


# ----------------------------------------------------------------------
# Public entry points
# ----------------------------------------------------------------------
def backfill():
    """Ingest the FULL monthly series from the latest NRB workbook. Returns count."""
    try:
        url = _fetch_latest_xlsx_url()
        if not url:
            logger.warning("No xlsx found on listing page — backfill skipped.")
            return 0
        logger.info("Backfilling interest rates from %s", url)
        parsed = parse_deposit_series(_download(url))
        if not parsed:
            logger.warning("No deposit series parsed — backfill skipped.")
            return 0
        merged = _merge(_read_rows(), parsed, f"{SOURCE_TAG} [{url.rsplit('/', 1)[-1]}]")
        _write_rows(merged)
        logger.info("Backfill merged %d monthly points (CSV now %d rows).", len(parsed), len(merged))
        return len(parsed)
    except Exception as e:
        logger.error("Interest-rate backfill failed: %s", e)
        return 0


def update():
    """Append only the newest month if not already recorded. Returns True if added."""
    try:
        url = _fetch_latest_xlsx_url()
        if not url:
            logger.info("Interest-rate update skipped — no xlsx found.")
            return False
        parsed = parse_deposit_series(_download(url))
        if not parsed:
            logger.info("Interest-rate update skipped — no value parsed.")
            return False

        latest = parsed[-1]
        existing = _read_rows()
        existing_dates = {r.get("date", "") for r in existing}
        if latest["date"] in existing_dates:
            logger.info("Interest-rate already recorded for %s.", latest["date"])
            return False

        merged = _merge(existing, [latest], f"{SOURCE_TAG} [{url.rsplit('/', 1)[-1]}]")
        _write_rows(merged)
        logger.info("Appended fixed-deposit rate %.2f%% for %s.", latest["rate"], latest["date"])
        return True
    except Exception as e:
        logger.error("Interest-rate update failed: %s", e)
        return False


def run_interest_rate_scrape():
    """Module-level entry used by the daily/month-end scraper."""
    return update()


if __name__ == "__main__":
    import argparse
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    parser = argparse.ArgumentParser(description="NRB bank interest-rate scraper")
    parser.add_argument("--backfill", action="store_true", help="Ingest full monthly history")
    args = parser.parse_args()

    if args.backfill:
        n = backfill()
        print(f"Backfilled {n} monthly points.")
    else:
        print("Appended new row." if update() else "No new row.")
